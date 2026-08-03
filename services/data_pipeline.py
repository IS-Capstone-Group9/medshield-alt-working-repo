"""MedShield sales and weather ingestion pipeline.

The module keeps raw source values for staging, produces canonical transaction
records for analytics, persists a local fallback, and writes to Supabase when a
service-role key is configured.
"""

from __future__ import annotations

import csv
import gzip
import hashlib
import io
import json
import math
import os
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from statistics import mean
from typing import Any, Iterable

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT_DIR / "data" / "medshield"
RAW_SALES_DIR = DATA_DIR / "raw" / "sales"
UPLOAD_DIR = DATA_DIR / "uploads"
PROCESSED_DIR = DATA_DIR / "processed"
SALES_DATASET_PATH = PROCESSED_DIR / "sales_transactions.json.gz"
SALES_STATUS_PATH = PROCESSED_DIR / "sales_dataset_status.json"
SALES_SNAPSHOT_PATH = PROCESSED_DIR / "dashboard_sales_snapshot.json"
WEATHER_DATASET_PATH = PROCESSED_DIR / "weather_signals.json"
DEFAULT_WORKBOOK_PATH = RAW_SALES_DIR / "Sales Report.xlsx"

load_dotenv(ROOT_DIR / ".env")

CANONICAL_FIELDS = [
    "area",
    "dr_number",
    "date_delivered",
    "product",
    "quantity",
    "unit_cost",
    "total_cost",
    "discount",
    "net_cost",
    "trade_price_unit",
    "total_trade_price",
    "net_income",
    "margin_pct",
]

HEADER_ALIASES = {
    "area": "area",
    "drnumber": "dr_number",
    "deliveryreceiptnumber": "dr_number",
    "datedelivered": "date_delivered",
    "deliverydate": "date_delivered",
    "date": "date_delivered",
    "product": "product",
    "qty": "quantity",
    "quantity": "quantity",
    "quantitysold": "quantity",
    "cp": "unit_cost",
    "unitcost": "unit_cost",
    "unitcostamount": "unit_cost",
    "totalcp": "total_cost",
    "totalcost": "total_cost",
    "totalcostamount": "total_cost",
    "disc": "discount",
    "discount": "discount",
    "discountamount": "discount",
    "netcp": "net_cost",
    "netcost": "net_cost",
    "netcostamount": "net_cost",
    "tpunit": "trade_price_unit",
    "tradepriceunit": "trade_price_unit",
    "tradepriceunitamount": "trade_price_unit",
    "totaltp": "total_trade_price",
    "totaltradeprice": "total_trade_price",
    "totaltradepriceamount": "total_trade_price",
    "netincome": "net_income",
    "netincomeamount": "net_income",
    "margin": "margin_pct",
    "marginpct": "margin_pct",
    "percentage": "margin_pct",
}

AREA_STANDARDIZATION = {
    "CAM NORTE": "Camarines Norte",
    "CAMARINES NORTE": "Camarines Norte",
    "CAM SUR": "Camarines Sur",
    "CAMARINES SUR": "Camarines Sur",
    "METRO MANILA": "Metro Manila",
    "NCR": "Metro Manila",
    "QUEZON PROVINCE": "Quezon",
}

# Only geographic territories receive coordinate-based weather features.
# Customer/channel labels found in the workbook remain valid sales dimensions,
# but they are intentionally excluded from meteorological joins.
AREA_COORDINATES = {
    "Batangas": (13.7565, 121.0583),
    "Camarines Norte": (14.1390, 122.7633),
    "Camarines Sur": (13.6218, 123.1948),
    "Cavite": (14.2456, 120.8786),
    "Laguna": (14.2691, 121.4113),
    "Marinduque": (13.4767, 121.9032),
    "Metro Manila": (14.5995, 120.9842),
    "Quezon": (13.9414, 121.6234),
    "Rizal": (14.6037, 121.3084),
}

NASA_POWER_URL = os.getenv(
    "NASA_POWER_DAILY_URL",
    "https://power.larc.nasa.gov/api/temporal/daily/point",
)
OPEN_METEO_ARCHIVE_URL = os.getenv(
    "OPEN_METEO_ARCHIVE_URL",
    "https://archive-api.open-meteo.com/v1/archive",
)


@dataclass
class SourceRow:
    workbook: str
    sheet: str
    row_number: int
    raw: dict[str, Any]


def _ensure_directories() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)


def _header_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _normalize_headers(values: Iterable[Any]) -> list[str]:
    headers: list[str] = []
    for value in values:
        if str(value or "").strip() == "%":
            headers.append("margin_pct")
            continue
        key = _header_key(value)
        headers.append(HEADER_ALIASES.get(key, key))
    return headers


def _clean_text(value: Any, *, upper: bool = False) -> str | None:
    if value is None:
        return None
    cleaned = re.sub(r"\s+", " ", str(value)).strip()
    if not cleaned:
        return None
    return cleaned.upper() if upper else cleaned


def _clean_area(value: Any) -> str | None:
    cleaned = _clean_text(value, upper=True)
    if not cleaned:
        return None
    return AREA_STANDARDIZATION.get(cleaned, cleaned.title())


def _clean_dr_number(value: Any) -> str | None:
    cleaned = _clean_text(value, upper=True)
    if not cleaned:
        return None
    cleaned = cleaned.replace("\u2010", "-").replace("\u2011", "-").replace("\u2012", "-")
    cleaned = cleaned.replace("\u2013", "-").replace("\u2014", "-").replace("\u2212", "-")
    cleaned = re.sub(r"^(DR\s*NUMBER|DR\s*NO\.?|D\.R\.|DR)\s*[:#-]?\s*", "DR-", cleaned)
    cleaned = re.sub(r"\s*-\s*", "-", cleaned)
    cleaned = re.sub(r"\s+", "", cleaned)
    cleaned = re.sub(r"-{2,}", "-", cleaned).strip("-")
    if cleaned.isdigit():
        cleaned = f"DR-{cleaned}"
    return cleaned or None


def _clean_number(value: Any, *, percent: bool = False) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip()
        is_percent = text.endswith("%")
        text = re.sub(r"[^0-9eE+\-.]", "", text)
        if not text:
            return None
        try:
            number = float(text)
        except ValueError:
            return None
        if is_percent:
            number /= 100
    if not math.isfinite(number):
        return None
    if percent and abs(number) > 2:
        number /= 100
    return round(number, 6)


def _clean_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    formats = (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d/%m/%Y",
        "%b %d, %Y",
        "%B %d, %Y",
    )
    for pattern in formats:
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _row_hash(row: dict[str, Any]) -> str:
    values = [str(row.get(field) if row.get(field) is not None else "") for field in CANONICAL_FIELDS]
    return hashlib.sha256("|".join(values).encode("utf-8")).hexdigest()


def _source_hash(source: SourceRow) -> str:
    payload = {
        "workbook": source.workbook,
        "sheet": source.sheet,
        "row": source.row_number,
        "raw": {key: str(value) if value is not None else None for key, value in source.raw.items()},
    }
    return hashlib.sha256(
        json.dumps(payload, sort_keys=True, ensure_ascii=True).encode("utf-8")
    ).hexdigest()


def _json_safe(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _detect_stage(headers: list[str], source_type: str) -> str:
    canonical_count = len(set(headers) & set(CANONICAL_FIELDS))
    quality_columns = {"quality_status", "quality_notes", "source_hash", "source_sheet"}
    if canonical_count == len(CANONICAL_FIELDS) and quality_columns.intersection(headers):
        return "cleaned"
    if canonical_count >= 10:
        return "raw_medshield" if source_type == "xlsx" else "raw_tabular"
    return "unknown"


def _read_xlsx(content: bytes, file_name: str) -> tuple[list[SourceRow], str, list[str]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    source_rows: list[SourceRow] = []
    detected_stages: list[str] = []
    detected_headers: list[str] = []

    for sheet in workbook.worksheets:
        header_row = None
        headers: list[str] = []
        for candidate_row in range(1, min(sheet.max_row, 20) + 1):
            values = [cell.value for cell in next(sheet.iter_rows(
                min_row=candidate_row,
                max_row=candidate_row,
            ))]
            normalized = _normalize_headers(values)
            if len(set(normalized) & set(CANONICAL_FIELDS)) >= 10:
                header_row = candidate_row
                headers = normalized
                detected_headers = headers
                detected_stages.append(_detect_stage(headers, "xlsx"))
                break
        if header_row is None:
            continue

        for row_number, values in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if not any(value is not None and str(value).strip() for value in values):
                continue
            raw = {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
                if header
            }
            source_rows.append(SourceRow(file_name, sheet.title, row_number, raw))

    if not source_rows:
        raise ValueError("No MedShield transaction table was found in the workbook.")
    stage = "cleaned" if detected_stages and all(item == "cleaned" for item in detected_stages) else "raw_medshield"
    return source_rows, stage, detected_headers


def _read_csv(content: bytes, file_name: str) -> tuple[list[SourceRow], str, list[str]]:
    text = content.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    if len(rows) < 2:
        raise ValueError("CSV must include a header and at least one transaction row.")
    header_index = None
    headers: list[str] = []
    for candidate_index, values in enumerate(rows[:20]):
        normalized = _normalize_headers(values)
        if len(set(normalized) & set(CANONICAL_FIELDS)) >= 10:
            header_index = candidate_index
            headers = normalized
            break
    if header_index is None:
        raise ValueError(
            "CSV does not contain the MedShield sales columns. "
            "Expected Area, DR Number, Date Delivered, Product, Qty, CP, Total CP, "
            "Disc, Net CP, TP/UNIT, TOTAL TP, Net Income, and %."
        )
    source_rows = []
    for row_number, values in enumerate(rows[header_index + 1:], start=header_index + 2):
        if not any(str(value).strip() for value in values):
            continue
        raw = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
            if header
        }
        sheet = str(raw.get("source_sheet") or _clean_date(raw.get("date_delivered")) or "CSV")
        source_rows.append(SourceRow(file_name, sheet, row_number, raw))
    return source_rows, _detect_stage(headers, "csv"), headers


def _read_source(content: bytes, file_name: str) -> tuple[list[SourceRow], str, list[str]]:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(content, file_name)
    if suffix == ".csv":
        return _read_csv(content, file_name)
    raise ValueError("Only .xlsx and .csv sales files are supported.")


def clean_sales_rows(
    source_rows: list[SourceRow],
    input_stage: str,
) -> tuple[list[dict[str, Any]], dict[str, Any], list[dict[str, Any]]]:
    cleaned_rows: list[dict[str, Any]] = []
    staging_rows: list[dict[str, Any]] = []
    quality_counts: Counter[str] = Counter()
    transformation_counts: Counter[str] = Counter()
    issue_counts: Counter[str] = Counter()
    seen_business_hashes: set[str] = set()

    for source in source_rows:
        raw = source.raw
        delivery_date = _clean_date(raw.get("date_delivered"))
        row = {
            "area": _clean_area(raw.get("area")),
            "dr_number": _clean_dr_number(raw.get("dr_number")),
            "date_delivered": delivery_date.isoformat() if delivery_date else None,
            "product": _clean_text(raw.get("product"), upper=True),
            "quantity": _clean_number(raw.get("quantity")),
            "unit_cost": _clean_number(raw.get("unit_cost")),
            "total_cost": _clean_number(raw.get("total_cost")),
            "discount": _clean_number(raw.get("discount")),
            "net_cost": _clean_number(raw.get("net_cost")),
            "trade_price_unit": _clean_number(raw.get("trade_price_unit")),
            "total_trade_price": _clean_number(raw.get("total_trade_price")),
            "net_income": _clean_number(raw.get("net_income")),
            "margin_pct": _clean_number(raw.get("margin_pct"), percent=True),
        }
        notes: list[str] = []
        transformations: list[str] = []

        if not row["area"]:
            notes.append("missing area")
            issue_counts["missing_area"] += 1
        if not delivery_date:
            notes.append("invalid or missing delivery date")
            issue_counts["invalid_or_missing_date"] += 1
        if not row["product"]:
            notes.append("missing product")
            issue_counts["missing_product"] += 1
        if raw.get("dr_number") is not None and row["dr_number"] != _clean_text(raw.get("dr_number"), upper=True):
            transformations.append("dr_number: standardized identifier format")

        source_year = int(source.sheet) if source.sheet.isdigit() and len(source.sheet) == 4 else None
        if delivery_date and source_year and delivery_date.year != source_year:
            notes.append(f"delivery year {delivery_date.year} differs from sheet {source_year}")

        for field in (
            "quantity",
            "unit_cost",
            "total_cost",
            "discount",
            "net_cost",
            "trade_price_unit",
            "total_trade_price",
            "net_income",
        ):
            if row[field] is None:
                row[field] = 0.0
                transformations.append(f"{field}: blank converted to 0")
                issue_counts["blank_numeric_values"] += 1
            elif field not in {"discount", "net_income"} and row[field] < 0:
                notes.append(f"{field} is negative")
                issue_counts["unexpected_negative_values"] += 1

        if row["margin_pct"] is None:
            net_cost = float(row["net_cost"] or 0)
            net_income = float(row["net_income"] or 0)
            row["margin_pct"] = round(net_income / net_cost, 6) if net_cost else 0.0
            transformations.append("margin_pct: derived from net income / net cost")
        elif row["margin_pct"] < -1 or row["margin_pct"] > 2:
            notes.append("margin percentage is outside the expected range")
            issue_counts["margin_anomalies"] += 1

        business_hash = _row_hash(row)
        duplicate = business_hash in seen_business_hashes
        if duplicate:
            notes.append("exact duplicate transaction")
            issue_counts["duplicate_rows"] += 1
        else:
            seen_business_hashes.add(business_hash)

        rejected = not row["area"] or not delivery_date or not row["product"]
        quality_status = "rejected" if rejected else ("warning" if notes else "valid")
        quality_counts[quality_status] += 1
        transformation_counts.update(transformations)

        source_hash = _source_hash(source)
        row.update({
            "year": delivery_date.year if delivery_date else None,
            "source_workbook": source.workbook,
            "source_sheet": source.sheet,
            "source_row_number": source.row_number,
            "source_hash": source_hash,
            "business_hash": business_hash,
            "quality_status": quality_status,
            "quality_notes": "; ".join(notes),
            "input_stage": input_stage,
            "standardization_applied": transformations,
            "duplicate": duplicate,
        })
        cleaned_rows.append(row)

        staging_rows.append({
            "source_workbook": source.workbook,
            "source_sheet": source.sheet,
            "source_row_number": source.row_number,
            "area_raw": _json_safe(raw.get("area")),
            "dr_number_raw": _json_safe(raw.get("dr_number")),
            "date_delivered_raw": _json_safe(raw.get("date_delivered")),
            "product_raw": _json_safe(raw.get("product")),
            "quantity_raw": _json_safe(raw.get("quantity")),
            "unit_cost_raw": _json_safe(raw.get("unit_cost")),
            "total_cost_raw": _json_safe(raw.get("total_cost")),
            "discount_raw": _json_safe(raw.get("discount")),
            "net_cost_raw": _json_safe(raw.get("net_cost")),
            "trade_price_unit_raw": _json_safe(raw.get("trade_price_unit")),
            "total_trade_price_raw": _json_safe(raw.get("total_trade_price")),
            "net_income_raw": _json_safe(raw.get("net_income")),
            "margin_pct_raw": _json_safe(raw.get("margin_pct")),
            "row_quality_status": quality_status,
            "row_quality_notes": "; ".join(notes),
            "source_hash": source_hash,
            "input_stage": input_stage,
            "standardization_applied": transformations,
        })

    accepted = [row for row in cleaned_rows if row["quality_status"] != "rejected"]
    dates = [row["date_delivered"] for row in accepted if row["date_delivered"]]
    years = Counter(str(row["year"]) for row in accepted if row["year"])
    quality_summary = {
        "input_stage": input_stage,
        "rows_extracted": len(cleaned_rows),
        "rows_accepted": len(accepted),
        "rows_rejected": quality_counts["rejected"],
        "rows_with_warnings": quality_counts["warning"],
        "duplicate_rows": sum(1 for row in cleaned_rows if row["duplicate"]),
        "valid_rows": quality_counts["valid"],
        "years": dict(sorted(years.items())),
        "source_period_start": min(dates) if dates else None,
        "source_period_end": max(dates) if dates else None,
        "standardizations": dict(transformation_counts),
        "issues": dict(issue_counts),
        "columns_received": CANONICAL_FIELDS,
    }
    return cleaned_rows, quality_summary, staging_rows


def build_dashboard_snapshot(rows: list[dict[str, Any]]) -> dict[str, Any]:
    accepted = [
        row for row in rows
        if row["quality_status"] != "rejected"
    ]
    monthly: dict[str, dict[str, float]] = defaultdict(lambda: {"revenue": 0.0, "income": 0.0})
    by_area: dict[str, dict[str, float]] = defaultdict(lambda: {"revenue": 0.0, "income": 0.0})
    by_product: dict[str, dict[str, float]] = defaultdict(
        lambda: {"revenue": 0.0, "income": 0.0, "qty": 0.0}
    )
    yearly: dict[str, dict[str, float]] = defaultdict(
        lambda: {"revenue": 0.0, "income": 0.0, "transactions": 0.0}
    )

    for row in accepted:
        period = str(row["date_delivered"])[:7]
        year = str(row["year"])
        revenue = float(row["total_trade_price"] or 0)
        income = float(row["net_income"] or 0)
        quantity = float(row["quantity"] or 0)
        monthly[period]["revenue"] += revenue
        monthly[period]["income"] += income
        by_area[str(row["area"])]["revenue"] += revenue
        by_area[str(row["area"])]["income"] += income
        by_product[str(row["product"])]["revenue"] += revenue
        by_product[str(row["product"])]["income"] += income
        by_product[str(row["product"])]["qty"] += quantity
        yearly[year]["revenue"] += revenue
        yearly[year]["income"] += income
        yearly[year]["transactions"] += 1

    total_revenue = sum(item["revenue"] for item in monthly.values())
    total_income = sum(item["income"] for item in monthly.values())
    ranked_products = sorted(by_product.items(), key=lambda item: item[1]["revenue"], reverse=True)
    cumulative = 0.0
    top_products = []
    for rank, (product, values) in enumerate(ranked_products, start=1):
        share = values["revenue"] / total_revenue if total_revenue else 0
        cumulative += share
        abc = "A" if cumulative <= 0.8 else ("B" if cumulative <= 0.95 else "C")
        top_products.append({
            "product": product,
            "revenue": round(values["revenue"], 2),
            "qty": round(values["qty"], 4),
            "income": round(values["income"], 2),
            "abc": abc,
            "pct_of_total": round(share * 100, 4),
            "rank": rank,
        })

    month_values: dict[int, list[float]] = defaultdict(list)
    for period, values in monthly.items():
        month_values[int(period[5:7])].append(values["revenue"])
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    return {
        "totals": {
            "total_revenue": round(total_revenue, 2),
            "total_income": round(total_income, 2),
            "total_transactions": len(accepted),
            "top_product": ranked_products[0][0] if ranked_products else "",
            "top_area": max(by_area, key=lambda key: by_area[key]["revenue"]) if by_area else "",
            "avg_margin": round(total_income / total_revenue, 6) if total_revenue else 0,
        },
        "monthly": [
            {"period": period, "revenue": round(values["revenue"], 2), "income": round(values["income"], 2)}
            for period, values in sorted(monthly.items())
        ],
        "by_area": [
            {"area": area, "revenue": round(values["revenue"], 2), "income": round(values["income"], 2)}
            for area, values in sorted(by_area.items(), key=lambda item: item[1]["revenue"], reverse=True)
        ],
        "top_products": top_products[:15],
        "year_summary": [
            {
                "year": year,
                "revenue": round(values["revenue"], 2),
                "income": round(values["income"], 2),
                "transactions": int(values["transactions"]),
            }
            for year, values in sorted(yearly.items())
        ],
        "seasonality": [
            {"month": month_names[index - 1], "avg_revenue": round(mean(month_values[index]), 2)}
            for index in range(1, 13)
            if month_values[index]
        ],
    }


def _uploaded_years(rows: list[dict[str, Any]]) -> list[int]:
    return sorted({
        int(row["year"])
        for row in rows
        if row.get("year") is not None and str(row.get("year")).isdigit()
    })


def _row_year(row: dict[str, Any]) -> int | None:
    value = row.get("year")
    if value is None or not str(value).isdigit():
        return None
    return int(value)


def _load_existing_sales_payload() -> dict[str, Any] | None:
    if not SALES_DATASET_PATH.exists():
        return None
    with gzip.open(SALES_DATASET_PATH, "rt", encoding="utf-8") as handle:
        return json.load(handle)


def _display_path(path_value: Path) -> str:
    try:
        return str(path_value.relative_to(ROOT_DIR))
    except ValueError:
        return str(path_value)


def _normalize_duplicate_flags(rows: list[dict[str, Any]]) -> None:
    seen_business_hashes: set[str] = set()
    for row in rows:
        business_hash = row.get("business_hash") or _row_hash(row)
        row["business_hash"] = business_hash
        duplicate = business_hash in seen_business_hashes
        row["duplicate"] = duplicate
        if row.get("quality_status") != "rejected":
            seen_business_hashes.add(business_hash)
        notes = [
            note.strip()
            for note in str(row.get("quality_notes") or "").split(";")
            if note.strip() and note.strip() != "exact duplicate transaction"
        ]
        if duplicate:
            notes.append("exact duplicate transaction")
            if row.get("quality_status") == "valid":
                row["quality_status"] = "warning"
        row["quality_notes"] = "; ".join(notes)


def _quality_summary_from_rows(
    rows: list[dict[str, Any]],
    *,
    input_stage: str,
    headers: list[str] | None = None,
    checksum: str | None = None,
    merge_strategy: str | None = None,
    merged_years: list[int] | None = None,
) -> dict[str, Any]:
    accepted = [row for row in rows if row.get("quality_status") != "rejected"]
    dates = [row["date_delivered"] for row in accepted if row.get("date_delivered")]
    quality_counts = Counter(str(row.get("quality_status") or "valid") for row in rows)
    transformation_counts: Counter[str] = Counter()
    issue_counts: Counter[str] = Counter()
    for row in rows:
        transformation_counts.update(row.get("standardization_applied") or [])
        for note in str(row.get("quality_notes") or "").split(";"):
            issue = note.strip().lower().replace(" ", "_")
            if issue:
                issue_counts[issue] += 1
    years = Counter(str(row["year"]) for row in accepted if row.get("year"))
    summary = {
        "input_stage": input_stage,
        "rows_extracted": len(rows),
        "rows_accepted": len(accepted),
        "rows_rejected": quality_counts["rejected"],
        "rows_with_warnings": quality_counts["warning"],
        "duplicate_rows": sum(1 for row in rows if row.get("duplicate")),
        "valid_rows": quality_counts["valid"],
        "years": dict(sorted(years.items())),
        "source_period_start": min(dates) if dates else None,
        "source_period_end": max(dates) if dates else None,
        "standardizations": dict(transformation_counts),
        "issues": dict(issue_counts),
        "columns_received": CANONICAL_FIELDS,
        "unique_products": len({row.get("product") for row in accepted if row.get("product")}),
        "unique_dr_numbers": len({row.get("dr_number") for row in accepted if row.get("dr_number")}),
        "sku_count": len({row.get("product") for row in accepted if row.get("product")}),
    }
    if headers is not None:
        summary["headers_detected"] = headers
    if checksum is not None:
        summary["checksum"] = checksum
    if merge_strategy is not None:
        summary["merge_strategy"] = merge_strategy
    if merged_years is not None:
        summary["merged_years"] = [str(year) for year in merged_years]
    return summary


def _merge_sales_history(
    incoming_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], str, list[int]]:
    existing_payload = _load_existing_sales_payload()
    years = _uploaded_years(incoming_rows)
    if not existing_payload or not existing_payload.get("rows"):
        merged_rows = list(incoming_rows)
        strategy = "created_new_history"
    elif years:
        year_set = set(years)
        merged_rows = [
            row for row in existing_payload["rows"]
            if _row_year(row) not in year_set
        ] + incoming_rows
        strategy = "replaced_uploaded_years"
    else:
        known_hashes = {
            row.get("source_hash")
            for row in existing_payload["rows"]
            if row.get("source_hash")
        }
        merged_rows = list(existing_payload["rows"]) + [
            row for row in incoming_rows
            if row.get("source_hash") not in known_hashes
        ]
        strategy = "appended_unknown_year_rows"
    merged_rows.sort(key=lambda row: (
        str(row.get("date_delivered") or "9999-12-31"),
        str(row.get("area") or ""),
        str(row.get("product") or ""),
        int(row.get("source_row_number") or 0),
    ))
    _normalize_duplicate_flags(merged_rows)
    return merged_rows, strategy, years


def _write_local_sales_dataset(
    rows: list[dict[str, Any]],
    quality_summary: dict[str, Any],
    file_name: str,
    checksum: str,
) -> dict[str, Any]:
    _ensure_directories()
    merged_rows, merge_strategy, merged_years = _merge_sales_history(rows)
    final_quality_summary = _quality_summary_from_rows(
        merged_rows,
        input_stage=quality_summary["input_stage"],
        headers=quality_summary.get("headers_detected"),
        checksum=checksum,
        merge_strategy=merge_strategy,
        merged_years=merged_years,
    )
    payload = {
        "metadata": {
            "dataset_name": "MedShield Pharmaceutical Sales",
            "source_file": file_name,
            "source_files": sorted({
                str(row.get("source_workbook"))
                for row in merged_rows
                if row.get("source_workbook")
            }),
            "checksum": checksum,
            "received_at": datetime.now(timezone.utc).isoformat(),
            "cleaning_status": "cleaned",
            "quality_summary": final_quality_summary,
            "canonical_columns": CANONICAL_FIELDS,
        },
        "rows": merged_rows,
    }
    with gzip.open(SALES_DATASET_PATH, "wt", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=True, separators=(",", ":"))
    SALES_STATUS_PATH.write_text(
        json.dumps(payload["metadata"], indent=2, ensure_ascii=True),
        encoding="utf-8",
    )
    SALES_SNAPSHOT_PATH.write_text(
        json.dumps(build_dashboard_snapshot(merged_rows), indent=2, ensure_ascii=True),
        encoding="utf-8",
    )
    return {
        "persisted": True,
        "path": _display_path(SALES_DATASET_PATH),
        "merge_strategy": merge_strategy,
        "years_replaced": [str(year) for year in merged_years],
        "total_rows": len(merged_rows),
    }


def _service_role_key() -> str:
    return os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()


def warehouse_write_enabled() -> bool:
    return bool(os.getenv("SUPABASE_URL", "").strip() and _service_role_key())


SUPABASE_SCHEMA_BY_RESOURCE = {
    "dim_date": "medshield_common",
    "dim_month": "medshield_common",
    "dim_area": "medshield_common",
    "dim_product": "medshield_common",
    "dim_product_alias": "medshield_common",
    "dim_source_system": "medshield_etl",
    "etl_pipeline_run": "medshield_etl",
    "etl_source_extract": "medshield_etl",
    "stg_sales_transactions": "medshield_sales",
    "fact_sales_transactions": "medshield_sales",
    "fact_monthly_sales": "medshield_sales",
    "fact_area_summary": "medshield_sales",
    "fact_product_summary": "medshield_sales",
    "fact_year_summary": "medshield_sales",
    "fact_seasonality": "medshield_sales",
    "fact_data_completeness": "medshield_sales",
    "stg_doh_historical": "medshield_external",
    "stg_pagasa_historical": "medshield_external",
    "stg_weather_api_observations": "medshield_external",
    "fact_disease_signal": "medshield_external",
    "fact_weather_signal": "medshield_external",
    "refresh_sales_aggregates": "medshield_sales",
}


def _supabase_resource_schema(resource: str) -> str | None:
    normalized = resource.removeprefix("rpc/")
    return SUPABASE_SCHEMA_BY_RESOURCE.get(normalized)


def _supabase_headers(prefer: str | None = None, schema: str | None = None) -> dict[str, str]:
    key = _service_role_key()
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }
    if schema:
        headers["Accept-Profile"] = schema
        headers["Content-Profile"] = schema
    if prefer:
        headers["Prefer"] = prefer
    return headers


def _supabase_request(
    method: str,
    resource: str,
    *,
    params: dict[str, Any] | None = None,
    payload: Any = None,
    prefer: str | None = None,
) -> Any:
    base_url = os.getenv("SUPABASE_URL", "").rstrip("/")
    schema = _supabase_resource_schema(resource)
    response = requests.request(
        method,
        f"{base_url}/rest/v1/{resource}",
        headers=_supabase_headers(prefer, schema),
        params=params,
        json=payload,
        timeout=60,
    )
    response.raise_for_status()
    return response.json() if response.content else None


def _batched(values: list[dict[str, Any]], size: int = 400) -> Iterable[list[dict[str, Any]]]:
    for start in range(0, len(values), size):
        yield values[start:start + size]


def _persist_sales_to_warehouse(
    rows: list[dict[str, Any]],
    staging_rows: list[dict[str, Any]],
    quality_summary: dict[str, Any],
    file_name: str,
    checksum: str,
) -> dict[str, Any]:
    if not warehouse_write_enabled():
        return {
            "configured": False,
            "persisted": False,
            "message": "Local fallback saved. Configure SUPABASE_SERVICE_ROLE_KEY for warehouse writes.",
        }

    run = _supabase_request(
        "POST",
        "etl_pipeline_run",
        payload=[{
            "pipeline_name": "medshield_sales_ingestion",
            "run_status": "running",
            "source_period_start": quality_summary["source_period_start"],
            "source_period_end": quality_summary["source_period_end"],
            "rows_extracted": quality_summary["rows_extracted"],
            "quality_summary": quality_summary,
        }],
        prefer="return=representation",
    )[0]
    run_key = run["pipeline_run_key"]

    try:
        source_rows = _supabase_request(
            "GET",
            "dim_source_system",
            params={"select": "source_system_key", "source_code": "eq.MEDSHIELD_XLSX", "limit": 1},
        )
        source_system_key = source_rows[0]["source_system_key"]
        for year in [int(item) for item in quality_summary.get("years", {})]:
            _supabase_request(
                "DELETE",
                "fact_sales_transactions",
                params={
                    "source_system_key": f"eq.{source_system_key}",
                    "and": f"(delivery_date_key.gte.{year}0101,delivery_date_key.lte.{year}1231)",
                },
            )

        dates = []
        areas = []
        products = []
        for row in rows:
            if row["quality_status"] == "rejected":
                continue
            delivery_date = datetime.strptime(row["date_delivered"], "%Y-%m-%d").date()
            dates.append({
                "date_key": int(delivery_date.strftime("%Y%m%d")),
                "calendar_date": delivery_date.isoformat(),
                "calendar_year": delivery_date.year,
                "calendar_quarter": ((delivery_date.month - 1) // 3) + 1,
                "calendar_month": delivery_date.month,
                "month_name": delivery_date.strftime("%B"),
                "month_short_name": delivery_date.strftime("%b"),
                "year_month": delivery_date.strftime("%Y-%m"),
                "day_of_month": delivery_date.day,
                "is_month_end": False,
            })
            areas.append({"area_name": row["area"], "area_group": "territory"})
            products.append({"product_name": row["product"]})

        unique_dates = list({item["date_key"]: item for item in dates}.values())
        unique_areas = list({item["area_name"]: item for item in areas}.values())
        unique_products = list({item["product_name"]: item for item in products}.values())
        for table, values, conflict in (
            ("dim_date", unique_dates, "date_key"),
            ("dim_area", unique_areas, "area_name"),
            ("dim_product", unique_products, "product_name"),
        ):
            for batch in _batched(values):
                _supabase_request(
                    "POST",
                    table,
                    params={"on_conflict": conflict},
                    payload=batch,
                    prefer="resolution=merge-duplicates",
                )

        area_map = {
            item["area_name"]: item["area_key"]
            for item in _supabase_request(
                "GET", "dim_area", params={"select": "area_key,area_name"}
            )
        }
        product_map = {
            item["product_name"]: item["product_key"]
            for item in _supabase_request(
                "GET", "dim_product", params={"select": "product_key,product_name"}
            )
        }

        for staging in staging_rows:
            staging["pipeline_run_key"] = run_key
        for batch in _batched(staging_rows):
            _supabase_request(
                "POST",
                "stg_sales_transactions",
                params={"on_conflict": "source_hash"},
                payload=batch,
                prefer="resolution=merge-duplicates",
            )

        facts = []
        for row in rows:
            if row["quality_status"] == "rejected":
                continue
            delivery_date = datetime.strptime(row["date_delivered"], "%Y-%m-%d").date()
            facts.append({
                "delivery_date_key": int(delivery_date.strftime("%Y%m%d")),
                "area_key": area_map[row["area"]],
                "product_key": product_map[row["product"]],
                "source_system_key": source_system_key,
                "pipeline_run_key": run_key,
                "dr_number": row["dr_number"],
                "quantity_sold": row["quantity"],
                "unit_cost_amount": row["unit_cost"],
                "total_cost_amount": row["total_cost"],
                "discount_amount": row["discount"],
                "net_cost_amount": row["net_cost"],
                "trade_price_unit_amount": row["trade_price_unit"],
                "total_trade_price_amount": row["total_trade_price"],
                "net_income_amount": row["net_income"],
                "margin_pct": row["margin_pct"],
                "source_workbook": row["source_workbook"],
                "source_sheet": row["source_sheet"],
                "source_row_number": row["source_row_number"],
                "source_hash": row["source_hash"],
            })
        for batch in _batched(facts):
            _supabase_request(
                "POST",
                "fact_sales_transactions",
                params={"on_conflict": "source_hash"},
                payload=batch,
                prefer="resolution=merge-duplicates",
            )

        snapshot_date = max(
            datetime.strptime(row["date_delivered"], "%Y-%m-%d").date()
            for row in rows
            if row["quality_status"] != "rejected"
        )
        _supabase_request(
            "POST",
            "rpc/refresh_sales_aggregates",
            payload={"p_snapshot_date_key": int(snapshot_date.strftime("%Y%m%d"))},
        )
        _supabase_request(
            "POST",
            "etl_source_extract",
            payload=[{
                "pipeline_run_key": run_key,
                "source_system_key": source_system_key,
                "source_name": file_name,
                "source_uri": "private-upload",
                "source_period_start": quality_summary["source_period_start"],
                "source_period_end": quality_summary["source_period_end"],
                "record_count": quality_summary["rows_extracted"],
                "checksum": checksum,
                "metadata_json": {"input_stage": quality_summary["input_stage"]},
            }],
        )
        _supabase_request(
            "PATCH",
            "etl_pipeline_run",
            params={"pipeline_run_key": f"eq.{run_key}"},
            payload={
                "run_status": "completed",
                "finished_at": datetime.now(timezone.utc).isoformat(),
                "rows_loaded": quality_summary["rows_accepted"],
                "rows_rejected": quality_summary["rows_rejected"],
                "quality_summary": quality_summary,
            },
        )
        return {"configured": True, "persisted": True, "pipeline_run_key": run_key}
    except Exception as error:
        _supabase_request(
            "PATCH",
            "etl_pipeline_run",
            params={"pipeline_run_key": f"eq.{run_key}"},
            payload={
                "run_status": "failed",
                "finished_at": datetime.now(timezone.utc).isoformat(),
                "rows_rejected": quality_summary["rows_rejected"],
                "quality_summary": quality_summary,
                "error_message": str(error)[:1000],
            },
        )
        raise


def ingest_sales_bytes(content: bytes, file_name: str, *, persist_raw: bool = True) -> dict[str, Any]:
    _ensure_directories()
    if not content:
        raise ValueError("Uploaded file is empty.")
    if len(content) > 30 * 1024 * 1024:
        raise ValueError("Uploaded file exceeds the 30 MB limit.")
    safe_name = re.sub(r"[^A-Za-z0-9._-]", "_", Path(file_name).name)
    checksum = hashlib.sha256(content).hexdigest()
    source_rows, input_stage, headers = _read_source(content, safe_name)
    cleaned_rows, quality_summary, staging_rows = clean_sales_rows(source_rows, input_stage)
    quality_summary["headers_detected"] = headers
    quality_summary["checksum"] = checksum

    if persist_raw:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        (UPLOAD_DIR / f"{timestamp}-{checksum[:12]}-{safe_name}").write_bytes(content)

    local_result = _write_local_sales_dataset(cleaned_rows, quality_summary, safe_name, checksum)
    try:
        warehouse = _persist_sales_to_warehouse(
            cleaned_rows,
            staging_rows,
            quality_summary,
            safe_name,
            checksum,
        )
    except Exception as error:
        warehouse = {
            "configured": warehouse_write_enabled(),
            "persisted": False,
            "message": f"Warehouse write failed after local cleaning was saved: {error}",
        }
    return {
        "dataset": {
            "file_name": safe_name,
            "input_stage": input_stage,
            "cleaning_status": "cleaned",
            "checksum": checksum,
        },
        "quality": quality_summary,
        "persistence": {
            "local": local_result,
            "warehouse": warehouse,
        },
    }


def ensure_baseline_sales_dataset() -> dict[str, Any]:
    if not DEFAULT_WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Baseline workbook not found: {DEFAULT_WORKBOOK_PATH}")
    needs_refresh = (
        not SALES_DATASET_PATH.exists()
        or SALES_DATASET_PATH.stat().st_mtime < DEFAULT_WORKBOOK_PATH.stat().st_mtime
    )
    if needs_refresh:
        return ingest_sales_bytes(
            DEFAULT_WORKBOOK_PATH.read_bytes(),
            DEFAULT_WORKBOOK_PATH.name,
            persist_raw=False,
        )
    return sales_dataset_status()


def _load_local_sales_payload() -> dict[str, Any]:
    ensure_baseline_sales_dataset()
    with gzip.open(SALES_DATASET_PATH, "rt", encoding="utf-8") as handle:
        return json.load(handle)


def sales_dataset_status() -> dict[str, Any]:
    if not SALES_STATUS_PATH.exists():
        return ensure_baseline_sales_dataset()
    return json.loads(SALES_STATUS_PATH.read_text(encoding="utf-8"))


def _filtered_sales_rows(
    payload: dict[str, Any],
    *,
    year: str | None = None,
    search: str = "",
    quality_status: str | None = None,
) -> list[dict[str, Any]]:
    rows = payload["rows"]
    normalized_search = search.strip().lower()
    filtered = []
    for row in rows:
        if year and year != "all" and str(row.get("year")) != year:
            continue
        if quality_status and quality_status != "all" and row.get("quality_status") != quality_status:
            continue
        if normalized_search:
            haystack = " ".join(
                str(row.get(field) or "")
                for field in ("area", "dr_number", "date_delivered", "product", "quality_notes")
            ).lower()
            if normalized_search not in haystack:
                continue
        filtered.append(row)
    return filtered


def sales_summary(
    *,
    year: str | None = None,
    search: str = "",
    quality_status: str | None = None,
) -> dict[str, Any]:
    payload = _load_local_sales_payload()
    filtered = _filtered_sales_rows(
        payload,
        year=year,
        search=search,
        quality_status=quality_status,
    )
    accepted = [row for row in filtered if row.get("quality_status") != "rejected"]
    numeric_fields = [
        "quantity",
        "unit_cost",
        "total_cost",
        "discount",
        "net_cost",
        "trade_price_unit",
        "total_trade_price",
        "net_income",
        "margin_pct",
    ]
    sums = {
        field: round(sum(float(row.get(field) or 0) for row in accepted), 4)
        for field in numeric_fields
    }
    averages = {
        field: round(sums[field] / len(accepted), 6) if accepted else 0
        for field in numeric_fields
    }
    area_revenue: Counter[str] = Counter()
    product_revenue: Counter[str] = Counter()
    for row in accepted:
        revenue = float(row.get("total_trade_price") or 0)
        if row.get("area"):
            area_revenue[str(row["area"])] += revenue
        if row.get("product"):
            product_revenue[str(row["product"])] += revenue
    return {
        "filters": {
            "year": year or "all",
            "search": search,
            "quality_status": quality_status or "all",
        },
        "counts": {
            "rows": len(filtered),
            "accepted_rows": len(accepted),
            "rejected_rows": sum(1 for row in filtered if row.get("quality_status") == "rejected"),
            "warning_rows": sum(1 for row in filtered if row.get("quality_status") == "warning"),
            "unique_products": len({row.get("product") for row in accepted if row.get("product")}),
            "sku_count": len({row.get("product") for row in accepted if row.get("product")}),
            "unique_dr_numbers": len({row.get("dr_number") for row in accepted if row.get("dr_number")}),
            "years": len({row.get("year") for row in accepted if row.get("year")}),
        },
        "sums": sums,
        "averages": averages,
        "top": {
            "product": product_revenue.most_common(1)[0][0] if product_revenue else "",
            "area": area_revenue.most_common(1)[0][0] if area_revenue else "",
        },
    }


def sales_page(
    *,
    year: str | None = None,
    page: int = 1,
    page_size: int = 25,
    search: str = "",
    quality_status: str | None = None,
) -> dict[str, Any]:
    payload = _load_local_sales_payload()
    filtered = _filtered_sales_rows(
        payload,
        year=year,
        search=search,
        quality_status=quality_status,
    )
    page_size = min(max(page_size, 10), 200)
    page_count = max(1, math.ceil(len(filtered) / page_size))
    page = min(max(page, 1), page_count)
    start = (page - 1) * page_size
    return {
        "metadata": payload["metadata"],
        "rows": filtered[start:start + page_size],
        "pagination": {
            "page": page,
            "page_size": page_size,
            "page_count": page_count if filtered else 0,
            "total_rows": len(filtered),
        },
        "filters": {
            "year": year or "all",
            "search": search,
            "quality_status": quality_status or "all",
        },
    }


def _severity_index(
    rainfall_mm: float,
    rainy_days: int,
    max_wind_kph: float,
    avg_temperature_c: float,
) -> float:
    rainfall_component = min(max(rainfall_mm, 0) / 500, 1) * 0.5
    rainy_days_component = min(max(rainy_days, 0) / 20, 1) * 0.2
    wind_component = min(max(max_wind_kph, 0) / 60, 1) * 0.2
    temperature_component = min(abs(avg_temperature_c - 27) / 10, 1) * 0.1
    return round(rainfall_component + rainy_days_component + wind_component + temperature_component, 4)


def _alert_level(severity: float) -> str:
    if severity >= 0.95:
        return "critical"
    if severity >= 0.80:
        return "warning"
    if severity >= 0.60:
        return "watch"
    return "normal"


def _monthly_weather_rows(
    area: str,
    provider: str,
    daily_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in daily_rows:
        grouped[str(row["date"])[:7]].append(row)
    monthly = []
    for period, rows in sorted(grouped.items()):
        rainfall = sum(float(row.get("precipitation_mm") or 0) for row in rows)
        rainy_days = sum(1 for row in rows if float(row.get("precipitation_mm") or 0) >= 1)
        temperatures = [float(row["temperature_c"]) for row in rows if row.get("temperature_c") is not None]
        humidities = [float(row["relative_humidity_pct"]) for row in rows if row.get("relative_humidity_pct") is not None]
        winds = [float(row["wind_speed_kph"]) for row in rows if row.get("wind_speed_kph") is not None]
        avg_temperature = mean(temperatures) if temperatures else 27.0
        avg_humidity = mean(humidities) if humidities else 0.0
        max_wind = max(winds) if winds else 0.0
        severity = _severity_index(rainfall, rainy_days, max_wind, avg_temperature)
        monthly.append({
            "period": period,
            "area": area,
            "provider": provider,
            "rainfall_mm": round(rainfall, 4),
            "rainy_days": rainy_days,
            "avg_temperature_c": round(avg_temperature, 4),
            "avg_relative_humidity_pct": round(avg_humidity, 4),
            "max_wind_speed_kph": round(max_wind, 4),
            "rainfall_severity_proxy": severity,
            "weather_alert_level": _alert_level(severity),
            "typhoon_flag": False,
            "high_wind_watch": max_wind >= 62,
            "weather_adjustment_factor": round(1 + min(severity, 1) * 0.2, 4),
            "source_period": period,
        })
    return monthly


def _daily_weather_rows(
    area: str,
    provider: str,
    daily_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows = []
    for row in daily_rows:
        rainfall = float(row.get("precipitation_mm") or 0)
        temperature = float(row.get("temperature_c") or 27)
        humidity = row.get("relative_humidity_pct")
        wind = float(row.get("wind_speed_kph") or 0)
        severity = _severity_index(rainfall, 1 if rainfall >= 1 else 0, wind, temperature)
        rows.append({
            "date": row["date"],
            "period": row["date"],
            "area": area,
            "provider": provider,
            "rainfall_mm": round(rainfall, 4),
            "rainy_day": rainfall >= 1,
            "temperature_c": round(temperature, 4),
            "relative_humidity_pct": round(float(humidity), 4) if humidity is not None else None,
            "wind_speed_kph": round(wind, 4),
            "rainfall_severity_proxy": severity,
            "weather_alert_level": _alert_level(severity),
            "typhoon_flag": False,
            "high_wind_watch": wind >= 62,
            "weather_adjustment_factor": round(1 + min(severity, 1) * 0.2, 4),
            "source_period": row["date"],
        })
    return rows


def _fetch_nasa_daily(latitude: float, longitude: float, start: date, end: date) -> list[dict[str, Any]]:
    response = requests.get(
        NASA_POWER_URL,
        params={
            "start": start.strftime("%Y%m%d"),
            "end": end.strftime("%Y%m%d"),
            "latitude": latitude,
            "longitude": longitude,
            "community": "AG",
            "parameters": "PRECTOTCORR,T2M,RH2M,WS10M",
            "format": "JSON",
            "units": "metric",
            "time-standard": "LST",
        },
        timeout=90,
    )
    response.raise_for_status()
    payload = response.json()
    parameters = payload["properties"]["parameter"]
    dates = sorted(parameters["PRECTOTCORR"])
    def valid(value: Any) -> float | None:
        if value in (None, -999):
            return None
        numeric = float(value)
        return None if numeric <= -998 else numeric

    return [{
        "date": datetime.strptime(day, "%Y%m%d").date().isoformat(),
        "precipitation_mm": valid(parameters["PRECTOTCORR"].get(day)),
        "temperature_c": valid(parameters["T2M"].get(day)),
        "relative_humidity_pct": valid(parameters["RH2M"].get(day)),
        "wind_speed_kph": (
            valid(parameters["WS10M"].get(day)) * 3.6
            if valid(parameters["WS10M"].get(day)) is not None
            else None
        ),
    } for day in dates if valid(parameters["PRECTOTCORR"].get(day)) is not None]


def _fetch_open_meteo_daily(latitude: float, longitude: float, start: date, end: date) -> list[dict[str, Any]]:
    response = requests.get(
        OPEN_METEO_ARCHIVE_URL,
        params={
            "latitude": latitude,
            "longitude": longitude,
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
            "daily": (
                "precipitation_sum,temperature_2m_mean,"
                "relative_humidity_2m_mean,wind_speed_10m_max"
            ),
            "timezone": "Asia/Manila",
            "wind_speed_unit": "kmh",
            "precipitation_unit": "mm",
        },
        timeout=90,
    )
    response.raise_for_status()
    daily = response.json()["daily"]
    return [{
        "date": day,
        "precipitation_mm": daily["precipitation_sum"][index],
        "temperature_c": daily["temperature_2m_mean"][index],
        "relative_humidity_pct": daily["relative_humidity_2m_mean"][index],
        "wind_speed_kph": daily["wind_speed_10m_max"][index],
    } for index, day in enumerate(daily["time"])]


def refresh_weather(
    *,
    start: date,
    end: date,
    areas: list[str] | None = None,
    provider: str = "nasa_power",
) -> dict[str, Any]:
    if end < start:
        raise ValueError("Weather end date must be on or after start date.")
    if (end - start).days > 3660:
        raise ValueError("Weather refresh is limited to ten years per request.")
    requested_areas = areas or sorted(AREA_COORDINATES)
    invalid_areas = [area for area in requested_areas if area not in AREA_COORDINATES]
    if invalid_areas:
        raise ValueError(f"No weather coordinates configured for: {', '.join(invalid_areas)}")
    provider = provider.strip().lower()
    if provider not in {"nasa_power", "open_meteo"}:
        raise ValueError("Weather provider must be nasa_power or open_meteo.")

    rows: list[dict[str, Any]] = []
    daily_validation_rows: list[dict[str, Any]] = []
    for area in requested_areas:
        latitude, longitude = AREA_COORDINATES[area]
        if provider == "nasa_power":
            daily = _fetch_nasa_daily(latitude, longitude, start, end)
        else:
            daily = _fetch_open_meteo_daily(latitude, longitude, start, end)
        daily_validation_rows.extend(_daily_weather_rows(area, provider, daily))
        rows.extend(_monthly_weather_rows(area, provider, daily))

    _ensure_directories()
    payload = {
        "metadata": {
            "provider": provider,
            "source_url": NASA_POWER_URL if provider == "nasa_power" else OPEN_METEO_ARCHIVE_URL,
            "refreshed_at": datetime.now(timezone.utc).isoformat(),
            "period_start": start.isoformat(),
            "period_end": end.isoformat(),
            "areas": requested_areas,
            "interpretation": (
                "NASA POWER and Open-Meteo produce a rainfall severity proxy, not official "
                "PAGASA RSI or a typhoon warning. The bounded uplift is a planning scenario "
                "and does not prove that weather caused pharmaceutical demand."
            ),
            "method": (
                "daily API observations retained for validation; monthly observed weather "
                "severity proxy used for bounded 0-20% planning scenario"
            ),
            "daily_rows_loaded": len(daily_validation_rows),
            "monthly_rows_loaded": len(rows),
        },
        "daily_rows": daily_validation_rows,
        "rows": rows,
    }
    WEATHER_DATASET_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    warehouse_result = _persist_weather_to_warehouse(rows, provider)
    return {
        "metadata": payload["metadata"],
        "rows_loaded": len(rows),
        "persistence": {
            "local": {"persisted": True, "path": str(WEATHER_DATASET_PATH.relative_to(ROOT_DIR))},
            "warehouse": warehouse_result,
        },
    }


def _persist_weather_to_warehouse(rows: list[dict[str, Any]], provider: str) -> dict[str, Any]:
    if not warehouse_write_enabled():
        return {"configured": False, "persisted": False}
    source_code = "NASA_POWER_DAILY" if provider == "nasa_power" else "OPEN_METEO_ARCHIVE"
    source = _supabase_request(
        "GET",
        "dim_source_system",
        params={"select": "source_system_key", "source_code": f"eq.{source_code}", "limit": 1},
    )
    if not source:
        raise RuntimeError(f"Source system {source_code} is missing. Apply migration 005 first.")
    source_key = source[0]["source_system_key"]
    unique_dates = {}
    for row in rows:
        period_date = datetime.strptime(f"{row['period']}-01", "%Y-%m-%d").date()
        unique_dates[int(period_date.strftime("%Y%m%d"))] = {
            "date_key": int(period_date.strftime("%Y%m%d")),
            "calendar_date": period_date.isoformat(),
            "calendar_year": period_date.year,
            "calendar_quarter": ((period_date.month - 1) // 3) + 1,
            "calendar_month": period_date.month,
            "month_name": period_date.strftime("%B"),
            "month_short_name": period_date.strftime("%b"),
            "year_month": period_date.strftime("%Y-%m"),
            "day_of_month": 1,
            "is_month_end": False,
        }
    _supabase_request(
        "POST",
        "dim_date",
        params={"on_conflict": "date_key"},
        payload=list(unique_dates.values()),
        prefer="resolution=merge-duplicates",
    )
    _supabase_request(
        "POST",
        "dim_area",
        params={"on_conflict": "area_name"},
        payload=[{"area_name": area, "area_group": "territory"} for area in sorted({row["area"] for row in rows})],
        prefer="resolution=merge-duplicates",
    )
    area_map = {
        item["area_name"]: item["area_key"]
        for item in _supabase_request("GET", "dim_area", params={"select": "area_key,area_name"})
    }
    facts = []
    for row in rows:
        period_date = datetime.strptime(f"{row['period']}-01", "%Y-%m-%d").date()
        facts.append({
            "period_date_key": int(period_date.strftime("%Y%m%d")),
            "area_key": area_map[row["area"]],
            "source_system_key": source_key,
            "rainfall_mm": row["rainfall_mm"],
            "rainfall_severity_index": row["rainfall_severity_proxy"],
            "rainfall_severity_proxy": row["rainfall_severity_proxy"],
            "rainy_days": row["rainy_days"],
            "avg_temperature_c": row["avg_temperature_c"],
            "avg_relative_humidity_pct": row["avg_relative_humidity_pct"],
            "max_wind_speed_kph": row["max_wind_speed_kph"],
            "weather_adjustment_factor": row["weather_adjustment_factor"],
            "typhoon_flag": row["typhoon_flag"],
            "high_wind_watch": row["high_wind_watch"],
            "weather_alert_level": row["weather_alert_level"],
            "source_period": row["source_period"],
            "provider_code": provider,
        })
    for batch in _batched(facts):
        _supabase_request(
            "POST",
            "fact_weather_signal",
            params={"on_conflict": "period_date_key,area_key,source_system_key"},
            payload=batch,
            prefer="resolution=merge-duplicates",
        )
    return {"configured": True, "persisted": True, "rows": len(facts)}


def weather_effects(
    *,
    year: str | None = None,
    area: str | None = None,
    grain: str = "monthly",
) -> dict[str, Any]:
    if not WEATHER_DATASET_PATH.exists():
        return {
            "metadata": {
                "status": "not_refreshed",
                "message": "Run a weather refresh to load NASA POWER or Open-Meteo data.",
            },
            "rows": [],
            "summary": [],
        }
    grain = grain.strip().lower()
    if grain not in {"daily", "monthly"}:
        raise ValueError("grain must be daily or monthly")
    payload = json.loads(WEATHER_DATASET_PATH.read_text(encoding="utf-8"))
    sales_rows = [
        row for row in _load_local_sales_payload()["rows"]
        if row["quality_status"] != "rejected"
    ]
    daily_sales: dict[tuple[str, str], float] = defaultdict(float)
    monthly_sales: dict[tuple[str, str], float] = defaultdict(float)
    for sale in sales_rows:
        daily_sales[(str(sale["date_delivered"]), str(sale["area"]))] += float(
            sale["net_cost"] or 0
        )
        monthly_sales[(str(sale["date_delivered"])[:7], str(sale["area"]))] += float(
            sale["net_cost"] or 0
        )
    rows = []
    by_area: dict[str, list[dict[str, Any]]] = defaultdict(list)
    daily_cache_missing = grain == "daily" and "daily_rows" not in payload
    source_rows = payload.get("daily_rows", []) if grain == "daily" else payload["rows"]
    for weather in source_rows:
        period = weather["date"] if grain == "daily" else weather["period"]
        if year and year != "all" and not period.startswith(year):
            continue
        if area and area != "all" and weather["area"] != area:
            continue
        revenue = (
            daily_sales.get((period, weather["area"]), 0.0)
            if grain == "daily"
            else monthly_sales.get((period, weather["area"]), 0.0)
        )
        item = {
            **weather,
            "period": period,
            "sales_revenue": revenue,
            "planning_demand_uplift_pct": round(
                (weather["weather_adjustment_factor"] - 1) * 100,
                2,
            ),
        }
        rows.append(item)
        by_area[weather["area"]].append(item)

    summary = []
    for area_name, area_rows in sorted(by_area.items()):
        severity = [float(row["rainfall_severity_proxy"]) for row in area_rows]
        revenue = [float(row["sales_revenue"]) for row in area_rows]
        correlation = None
        if len(severity) >= 3 and len(set(severity)) > 1 and len(set(revenue)) > 1:
            severity_mean = mean(severity)
            revenue_mean = mean(revenue)
            numerator = sum(
                (left - severity_mean) * (right - revenue_mean)
                for left, right in zip(severity, revenue)
            )
            denominator = math.sqrt(
                sum((value - severity_mean) ** 2 for value in severity)
                * sum((value - revenue_mean) ** 2 for value in revenue)
            )
            correlation = round(numerator / denominator, 4) if denominator else None
        summary.append({
            "area": area_name,
            "periods": len(area_rows),
            "months": len(area_rows) if grain == "monthly" else len({str(row["period"])[:7] for row in area_rows}),
            "days": len(area_rows) if grain == "daily" else None,
            "sales_matched_periods": sum(1 for row in area_rows if float(row.get("sales_revenue") or 0) > 0),
            "avg_rainfall_severity_proxy": round(mean(severity), 4) if severity else 0,
            "max_planning_uplift_pct": round(
                max(row["planning_demand_uplift_pct"] for row in area_rows),
                2,
            ) if area_rows else 0,
            "rainfall_revenue_correlation": correlation,
            "interpretation": (
                "Positive values indicate that rainfall severity and sales moved together; "
                "negative values indicate opposite movement. This is association, not causation."
            ),
        })
    metadata = {
        **payload["metadata"],
        "grain": grain,
        "rows_returned": len(rows),
        "areas_returned": sorted({row["area"] for row in rows}),
        "sales_matched_rows": sum(1 for row in rows if float(row.get("sales_revenue") or 0) > 0),
    }
    if daily_cache_missing:
        metadata["message"] = (
            "The current weather cache only has monthly rows. Click Refresh weather to load daily API validation rows."
        )
    elif not rows:
        metadata["message"] = (
            "No weather rows match this selection. Refresh all territories or select a loaded territory/year."
        )
    elif metadata["sales_matched_rows"] == 0:
        metadata["message"] = (
            "Weather rows loaded, but no same-period sales rows match this territory. "
            "This can happen when the sales file uses channel labels or another area naming convention."
        )
    return {"metadata": metadata, "rows": rows, "summary": summary}
