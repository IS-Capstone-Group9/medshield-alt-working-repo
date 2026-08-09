import openpyxl
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
EXCEL_FILE = ROOT / "Medshield (Keith San Miguel).xlsx"
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

# Build product signature dictionary from 2018-2025 clean sheets
signature_map = {}
signature_counts = defaultdict(set)

for sheet_name in ["sales2017", "sales2018", "sales2019", "sales2020", "sales2021", "sales2022", "sales2023", "sales2024", "sales2025"]:
    if sheet_name not in wb.sheetnames:
        continue
    sheet = wb[sheet_name]
    header_idx = None
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        vals = [str(c).strip().lower() if c is not None else "" for c in row]
        if "product" in vals or any("product" in c for c in vals):
            header_idx = r_idx
            break
    if not header_idx:
        continue
    
    for row in sheet.iter_rows(min_row=header_idx+1, values_only=True):
        if not any(row):
            continue
        # Check if product is clean
        prod = str(row[3] if len(row) > 3 else "").strip()
        if not prod or prod == "#REF!" or prod == "None":
            continue
        
        # Extract numeric cost/price features
        try:
            # Let's inspect typical position of CP and TP
            cp = float(row[5]) if len(row) > 5 and row[5] is not None else None
            tp = float(row[9]) if len(row) > 9 and row[9] is not None else None
            if cp is not None and cp > 0:
                key = (round(cp, 2), round(tp, 2) if tp else 0.0)
                signature_counts[key].add(prod)
        except (ValueError, TypeError):
            continue

print(f"Total unique cost/price signatures built: {len(signature_counts)}")
unique_signatures = {k: list(v)[0] for k, v in signature_counts.items() if len(v) == 1}
print(f"Unique unambiguous signatures: {len(unique_signatures)}")

# Test matching on 2017 #REF! rows
sheet17 = wb["sales2017"]
rows17 = list(sheet17.iter_rows(values_only=True))

matched_count = 0
unmatched_count = 0

for r in rows17[6:]:
    prod = str(r[3] if len(r) > 3 else "").strip()
    if prod == "#REF!":
        try:
            cp = float(r[5]) if len(r) > 5 and r[5] is not None else None
            tp = float(r[10]) if len(r) > 10 and r[10] is not None else None
            key = (round(cp, 2) if cp else 0.0, round(tp, 2) if tp else 0.0)
            if key in unique_signatures:
                matched_count += 1
            else:
                unmatched_count += 1
        except (ValueError, TypeError):
            unmatched_count += 1

print(f"2017 #REF! rows matched: {matched_count}, remaining unmatched: {unmatched_count}")
