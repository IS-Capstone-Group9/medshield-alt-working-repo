import openpyxl
import pandas as pd
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
EXCEL_FILE = ROOT / "Medshield (Keith San Miguel).xlsx"
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

print("=== INSPECTING 2017 SHEET ===")
sheet17 = wb["sales2017"]
rows17 = list(sheet17.iter_rows(values_only=True))
print("Total rows in sales2017:", len(rows17))
header17 = rows17[4]
print("Header row 5:", header17)
ref_count_17 = 0
valid_prod_count_17 = 0
for idx, r in enumerate(rows17[6:], start=7):
    prod = r[3]
    if str(prod).strip() == "#REF!":
        ref_count_17 += 1
    elif prod and str(prod).strip():
        valid_prod_count_17 += 1
print(f"2017 #REF! count: {ref_count_17}, Valid product count: {valid_prod_count_17}")
if valid_prod_count_17 > 0:
    print("Sample valid 2017 rows:")
    count = 0
    for r in rows17[6:]:
        if r[3] and str(r[3]).strip() != "#REF!":
            print("  ", r)
            count += 1
            if count >= 5:
                break

print("\n=== INSPECTING 2018 SHEET ===")
sheet18 = wb["sales2018"]
rows18 = list(sheet18.iter_rows(values_only=True))
print("Total rows in sales2018:", len(rows18))
header18 = rows18[4]
print("Header row 5:", header18)
ref_count_18 = 0
valid_prod_count_18 = 0
for r in rows18[5:]:
    prod = r[3]
    if str(prod).strip() == "#REF!":
        ref_count_18 += 1
    elif prod and str(prod).strip():
        valid_prod_count_18 += 1
print(f"2018 #REF! count: {ref_count_18}, Valid product count: {valid_prod_count_18}")

print("\n=== INSPECTING 2019 SHEET ===")
sheet19 = wb["sales2019"]
rows19 = list(sheet19.iter_rows(values_only=True))
print("Total rows in sales2019:", len(rows19))
header19 = rows19[4]
print("Header row 5:", header19)
ref_count_19 = 0
valid_prod_count_19 = 0
for r in rows19[6:]:
    prod = r[3]
    if str(prod).strip() == "#REF!":
        ref_count_19 += 1
    elif prod and str(prod).strip():
        valid_prod_count_19 += 1
print(f"2019 #REF! count: {ref_count_19}, Valid product count: {valid_prod_count_19}")
