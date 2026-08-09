import csv
from datetime import datetime, date
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
EXCEL_FILE = ROOT / "Medshield (Keith San Miguel).xlsx"
TARGET_DIR = ROOT / "data" / "medshield" / "raw" / "sales" / "yearly_csv"
TARGET_DIR.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

years_mapping = {
    "sales2017": "2017s.csv",
    "sales2018": "2018s.csv",
    "sales2019": "2019s.csv",
    "sales2020": "2020s.csv",
}

for sheet_name, csv_name in years_mapping.items():
    if sheet_name not in wb.sheetnames:
        print(f"Warning: {sheet_name} not found in workbook")
        continue
    
    sheet = wb[sheet_name]
    target_path = TARGET_DIR / csv_name
    
    rows_written = 0
    with target_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for row in sheet.iter_rows(values_only=True):
            # Format dates to string
            formatted_row = []
            for val in row:
                if isinstance(val, (datetime, date)):
                    formatted_row.append(val.strftime("%Y-%m-%d"))
                elif val is None:
                    formatted_row.append("")
                else:
                    formatted_row.append(str(val))
            writer.writerow(formatted_row)
            rows_written += 1
            
    print(f"Exported {sheet_name} -> {target_path.relative_to(ROOT)} ({rows_written} rows)")
