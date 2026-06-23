import argparse
import csv
import gzip
import json
import re
import subprocess
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW_WORKBOOK = ROOT / "Sales Report.xlsx"
ALLOCATED_DATASET = ROOT / "data" / "medshield" / "processed" / "sales_transactions_area_allocated.json.gz"
LAYER_DIR = ROOT / "outputs" / "sales_data_layers"

YEARS = ["2021", "2022", "2023", "2024", "2025"]
ADDITIVE_FIELDS = [
    "quantity",
    "total_cost",
    "discount",
    "net_cost",
    "total_trade_price",
    "net_income",
]
PRIORITY_FIELDS = [
    "data_layer",
    "date_delivered",
    "date_is_estimated",
    "date_estimation_method",
    "original_date_delivered",
    "nearest_dated_source_row_number",
    "date_source_row_distance",
    "year",
    "month",
    "day",
    "dr_number",
    "area",
    "product",
    "quantity",
    "unit_cost",
    "total_cost",
    "discount",
    "net_cost",
    "trade_price_unit",
    "total_trade_price",
    "net_income",
    "margin_pct",
    "quality_status",
    "quality_notes",
    "allocation_status",
    "estimated",
    "original_area",
    "original_product",
    "allocation_method",
    "allocation_confidence",
    "allocation_profile_level",
    "allocation_weight",
    "allocation_child_number",
    "allocation_child_count",
    "embedded_area_amount",
    "allocated_embedded_area_amount",
    "source_workbook",
    "source_sheet",
    "source_row_number",
    "source_hash",
    "business_hash",
    "parent_source_hash",
    "parent_business_hash",
    "input_stage",
    "standardization_applied",
    "duplicate",
    "allocation_profile_rows",
    "allocation_profile_products",
    "sales_acceptance_status",
    "sales_rejection_reason",
]


def is_blank(value):
    return value is None or str(value).strip() == ""


def is_valid_date(value):
    return bool(re.match(r"^\d{4}-\d{2}-\d{2}$", str(value or "")))


def workbook_year(row):
    text = f"{row.get('source_workbook', '')} {row.get('source_sheet', '')}"
    match = re.search(r"(20\d{2})", text)
    return match.group(1) if match else ""


def row_year(row):
    if is_valid_date(row.get("date_delivered")):
        return str(row["date_delivered"])[:4]
    return str(row.get("year") or workbook_year(row) or "")


def load_json_gz(path):
    with gzip.open(path, "rt", encoding="utf-8") as handle:
        return json.load(handle)


def load_json_gz_from_git(ref, repo_path):
    raw = subprocess.check_output(
        ["git", "show", f"{ref}:{repo_path}"],
        cwd=ROOT,
    )
    return json.loads(gzip.decompress(raw).decode("utf-8"))


def write_csv(path, rows, fields=None):
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = list(rows)
    if fields is None:
        fields = build_fields(rows)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def build_fields(rows):
    fields = []
    seen = set()
    for field in PRIORITY_FIELDS:
        fields.append(field)
        seen.add(field)
    for row in rows:
        for key in row.keys():
            if key not in seen:
                fields.append(key)
                seen.add(key)
    return fields


def clean_cell(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return value


def export_raw_workbook():
    raw_dir = LAYER_DIR / "raw_sales_report"
    workbook = load_workbook(RAW_WORKBOOK, read_only=True, data_only=True)
    summaries = {}

    for sheet in workbook.worksheets:
        year_match = re.search(r"(20\d{2})", sheet.title)
        if not year_match:
            continue
        year = year_match.group(1)
        rows = []
        header_row_number = None
        header_values = None
        for row_number, row_values in enumerate(sheet.iter_rows(values_only=True), start=1):
            normalized = [str(value).strip().lower() if value is not None else "" for value in row_values]
            if "area" in normalized and "product" in normalized:
                header_row_number = row_number
                header_values = row_values
                break
        if header_row_number is None or header_values is None:
            raise ValueError(f"Could not find sales header row for sheet {sheet.title}")
        header_values = header_values[:13]
        headers = [
            str(value).strip() if not is_blank(value) else f"column_{index + 1}"
            for index, value in enumerate(header_values)
        ]
        for source_row_number, row_values in enumerate(
            sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            if not any(not is_blank(value) for value in row_values):
                continue
            row = {
                "data_layer": "raw_sales_report",
                "source_workbook": RAW_WORKBOOK.name,
                "source_sheet": sheet.title,
                "source_row_number": source_row_number,
                "source_header_row_number": header_row_number,
            }
            for header, value in zip(headers, row_values[:13]):
                row[header] = clean_cell(value)
            rows.append(row)

        raw_fields = [
            "data_layer",
            "source_workbook",
            "source_sheet",
            "source_row_number",
            "source_header_row_number",
            *headers,
        ]
        write_csv(raw_dir / f"sales_report_{year}_raw.csv", rows, raw_fields)
        summaries[year] = {
            "rows": len(rows),
            "columns": len(headers),
            "header_row_number": header_row_number,
        }

    return summaries


def get_base_rejection_reasons(row):
    reasons = []
    if is_blank(row.get("product")):
        reasons.append("missing_product")
    if is_blank(row.get("area")):
        reasons.append("missing_area")
    if not is_valid_date(row.get("date_delivered")):
        reasons.append("missing_or_invalid_date_delivered")
    if not (float(row.get("quantity") or 0) > 0):
        reasons.append("non_positive_quantity")
    if not (float(row.get("total_trade_price") or 0) > 0):
        reasons.append("non_positive_total_trade_price")
    return reasons


def build_date_estimator(rows):
    dated_by_workbook = defaultdict(list)
    for row in rows:
        if not is_valid_date(row.get("date_delivered")):
            continue
        workbook = row.get("source_workbook") or workbook_year(row) or str(row["date_delivered"])[:4]
        dated_by_workbook[workbook].append(
            {
                "source_row_number": int(float(row.get("source_row_number") or 0)),
                "date_delivered": str(row["date_delivered"]),
            }
        )

    for candidates in dated_by_workbook.values():
        candidates.sort(key=lambda item: item["source_row_number"])

    def estimate(row):
        workbook = row.get("source_workbook") or workbook_year(row)
        candidates = dated_by_workbook.get(workbook, [])
        source_row_number = int(float(row.get("source_row_number") or 0))
        if candidates:
            best = min(
                candidates,
                key=lambda item: abs(item["source_row_number"] - source_row_number),
            )
            return {
                "date": best["date_delivered"],
                "method": "estimated_nearest_dated_row_same_workbook",
                "nearest": best["source_row_number"],
                "distance": abs(best["source_row_number"] - source_row_number),
            }
        year = workbook_year(row) or "2025"
        return {
            "date": f"{year}-07-01",
            "method": "estimated_midyear_fallback_from_source_workbook",
            "nearest": "",
            "distance": "",
        }

    return estimate


def prepare_semi_raw_rows(rows):
    estimate_date = build_date_estimator(rows)
    prepared = []

    for original in rows:
        row = dict(original)
        row["data_layer"] = "semi_raw_backwards_approx"
        row["original_date_delivered"] = row.get("date_delivered") or ""

        if is_valid_date(row.get("date_delivered")):
            row["date_is_estimated"] = "false"
            row["date_estimation_method"] = "source_date_delivered"
            row["nearest_dated_source_row_number"] = ""
            row["date_source_row_distance"] = ""
        else:
            estimate = estimate_date(row)
            row["date_delivered"] = estimate["date"]
            row["date_is_estimated"] = "true"
            row["date_estimation_method"] = estimate["method"]
            row["nearest_dated_source_row_number"] = estimate["nearest"]
            row["date_source_row_distance"] = estimate["distance"]

        if is_valid_date(row.get("date_delivered")):
            row["year"] = str(row["date_delivered"])[:4]
            row["month"] = str(row["date_delivered"])[5:7]
            row["day"] = str(row["date_delivered"])[8:10]
        else:
            row["year"] = row_year(row)
            row["month"] = ""
            row["day"] = ""

        rejection_reasons = get_base_rejection_reasons(row)
        if rejection_reasons:
            row["sales_acceptance_status"] = "pending_cleaning_excluded"
            row["sales_rejection_reason"] = "; ".join(rejection_reasons)
        else:
            row["sales_acceptance_status"] = "eligible_for_clean_sales"
            row["sales_rejection_reason"] = ""

        prepared.append(row)

    return prepared


def prepare_clean_rows(rows):
    prepared = []
    for original in rows:
        row = dict(original)
        row["data_layer"] = "cleaned_sales"
        row["sales_acceptance_status"] = row.get("sales_acceptance_status") or "accepted_clean_sales"
        row["sales_rejection_reason"] = ""
        if is_valid_date(row.get("date_delivered")):
            row["year"] = str(row["date_delivered"])[:4]
            row["month"] = str(row["date_delivered"])[5:7]
            row["day"] = str(row["date_delivered"])[8:10]
        prepared.append(row)
    return prepared


def summarize_rows(rows):
    summary = {
        "rows": len(rows),
        "by_year": dict(Counter(row_year(row) for row in rows)),
        "empty_product": sum(1 for row in rows if is_blank(row.get("product"))),
        "empty_area": sum(1 for row in rows if is_blank(row.get("area"))),
        "bad_date": sum(1 for row in rows if not is_valid_date(row.get("date_delivered"))),
        "estimated_dates": sum(1 for row in rows if str(row.get("date_is_estimated")) == "true"),
        "estimated_backward_allocation_rows": sum(
            1 for row in rows if row.get("allocation_status") == "estimated_backward_allocation"
        ),
        "pending_cleaning_excluded": sum(
            1 for row in rows if row.get("sales_acceptance_status") == "pending_cleaning_excluded"
        ),
    }
    for field in ADDITIVE_FIELDS:
        summary[field] = round(sum(float(row.get(field) or 0) for row in rows), 6)
    return summary


def export_layer(name, rows):
    layer_dir = LAYER_DIR / name
    fields = build_fields(rows)
    write_csv(layer_dir / f"{name}_all_years.csv", rows, fields)
    by_year = defaultdict(list)
    for row in rows:
        by_year[row_year(row)].append(row)
    for year in YEARS:
        write_csv(layer_dir / f"{name}_{year}.csv", by_year.get(year, []), fields)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--semi-ref",
        default="98ebe69",
        help="Git ref containing the full backward-approximated dataset before clean-sales filtering.",
    )
    args = parser.parse_args()

    raw_summary = export_raw_workbook()

    semi_payload = load_json_gz_from_git(
        args.semi_ref,
        "data/medshield/processed/sales_transactions_area_allocated.json.gz",
    )
    semi_rows = prepare_semi_raw_rows(semi_payload["rows"])
    semi_rows.sort(
        key=lambda row: (
            str(row.get("date_delivered") or ""),
            str(row.get("area") or ""),
            str(row.get("product") or ""),
            int(float(row.get("source_row_number") or 0)),
        )
    )

    clean_payload = load_json_gz(ALLOCATED_DATASET)
    clean_rows = prepare_clean_rows(clean_payload["rows"])
    clean_rows.sort(
        key=lambda row: (
            str(row.get("date_delivered") or ""),
            str(row.get("area") or ""),
            str(row.get("product") or ""),
            int(float(row.get("source_row_number") or 0)),
        )
    )

    export_layer("semi_raw_backwards_approx", semi_rows)
    export_layer("cleaned_sales", clean_rows)

    summary = {
        "flow": [
            "raw_sales_report",
            "semi_raw_backwards_approx",
            "cleaned_sales",
        ],
        "raw_sales_report": raw_summary,
        "semi_raw_backwards_approx": summarize_rows(semi_rows),
        "cleaned_sales": summarize_rows(clean_rows),
        "notes": [
            "Raw sales report exports are sheet data copied from Sales Report.xlsx without cleaning.",
            "Semi-raw backward approximation keeps all allocated rows, repairs missing dates for traceability, and flags rows that fail clean-sales rules.",
            "Cleaned sales contains only accepted rows with product, area, valid date, positive quantity, and positive total trade price.",
            "Contract-name rows with # are not treated as products in the cleaned analytical layer; they are represented through estimated child rows when allocation was possible.",
        ],
    }
    (LAYER_DIR / "sales_data_layer_summary.json").write_text(
        json.dumps(summary, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
